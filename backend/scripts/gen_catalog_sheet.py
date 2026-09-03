"""Build the uploadable catalog workbook: docs/kirana-catalog.xlsx.

A script rather than a checked-in binary nobody can diff. The product list
below is the actual source of truth; regenerate with

    cd backend && uv run python scripts/gen_catalog_sheet.py

THE FIRST SHEET IS THE ONLY ONE PARSED. `catalog_service._read_rows` reads
`wb[wb.sheetnames[0]]`, so the second sheet can say anything without any risk
of a stray column being mistaken for a price. That is why the guidance lives
there instead of in a comment column next to the data.

WHY THE COSTS LOOK GENEROUS. They are chosen so the shop can actually haggle,
and that is a demo decision, not a claim about Indian retail. A kirana's real
margin on atta or oil is 3-8%, and the gate's margin floor is measured on the
sale price, so with a 12% floor an item carrying an 14% margin can give away
2% and no more -- which is exactly the "I cannot go below cost on this one"
the live shop was producing. The relationship is

    max_discount = 1 - (1 - margin) / (1 - floor)

so at a 12% floor a 20% margin buys 9.1% of headroom and a 25% margin buys
14.8%. Most rows here sit at 20-27%. Substitute real costs and the gate will
narrow the negotiation to match, honestly and immediately.

THREE ROWS ARE DELIBERATELY THIN. Sugar (10.4%) and salt (10.7%) sit BELOW a
12% floor and are refused outright; butter (13.8%) sits just above it and can
give away 2%. That is not an oversight -- a catalogue where everything is
discountable never shows the gate saying no, and the refusal is the product.

THE SIX ORIGINAL SKUS ARE PRESERVED. ATTA5, RICE5, DAL1K, OIL1L, SUGAR1 and
TEA250 keep their codes and names, because `shelf_items.sku` and
`slots.bound_sku` reference them by code and a Replace import deactivates
anything absent. Renaming them would quietly empty existing shelves.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.bounds import max_discount_for_margin  # noqa: E402

# --------------------------------------------------------------- the shop ---
# (sku, name, unit, price_rupees, cost_rupees, category)
PRODUCTS: list[tuple[str, str, str, str, str, str]] = [
    # --- atta, flour and rice -----------------------------------------------
    ("ATTA5",      "Aashirvaad Whole Wheat Atta 5kg",   "bag",    "285.00", "228.00", "Atta & Flour"),
    ("ATTA10",     "Aashirvaad Whole Wheat Atta 10kg",  "bag",    "545.00", "436.00", "Atta & Flour"),
    ("MAIDA1",     "Maida Refined Flour 1kg",           "pack",   "52.00",  "42.00",  "Atta & Flour"),
    ("BESAN1",     "Besan Gram Flour 1kg",              "pack",   "96.00",  "76.00",  "Atta & Flour"),
    ("SUJI1",      "Sooji Rava 1kg",                    "pack",   "58.00",  "46.00",  "Atta & Flour"),
    ("RICE5",      "India Gate Basmati Rice 5kg",       "bag",    "620.00", "496.00", "Rice"),
    ("RICE1",      "Sona Masoori Rice 1kg",             "pack",   "78.00",  "62.00",  "Rice"),
    ("POHA1",      "Poha Thick 1kg",                    "pack",   "62.00",  "49.00",  "Rice"),

    # --- dals and pulses ----------------------------------------------------
    ("DAL1K",      "Toor Dal 1kg",                      "pack",   "175.00", "140.00", "Dals & Pulses"),
    ("DALMOONG",   "Moong Dal 1kg",                     "pack",   "138.00", "110.00", "Dals & Pulses"),
    ("DALCHANA",   "Chana Dal 1kg",                     "pack",   "92.00",  "73.00",  "Dals & Pulses"),
    ("DALURAD",    "Urad Dal 1kg",                      "pack",   "145.00", "116.00", "Dals & Pulses"),
    ("DALMASOOR",  "Masoor Dal 1kg",                    "pack",   "118.00", "94.00",  "Dals & Pulses"),
    ("RAJMA1",     "Rajma Kidney Beans 1kg",            "pack",   "165.00", "132.00", "Dals & Pulses"),
    ("CHOLE1",     "Kabuli Chana 1kg",                  "pack",   "135.00", "108.00", "Dals & Pulses"),

    # --- oils and ghee ------------------------------------------------------
    ("OIL1L",      "Fortune Sunflower Oil 1L",          "bottle", "145.00", "116.00", "Oils & Ghee"),
    ("OIL5L",      "Fortune Sunflower Oil 5L",          "can",    "690.00", "552.00", "Oils & Ghee"),
    ("MUSTOIL1",   "Mustard Oil 1L",                    "bottle", "168.00", "134.00", "Oils & Ghee"),
    ("GHEE500",    "Amul Pure Ghee 500ml",              "jar",    "315.00", "252.00", "Oils & Ghee"),
    ("GHEE1L",     "Amul Pure Ghee 1L",                 "tin",    "620.00", "496.00", "Oils & Ghee"),

    # --- sugar, salt, sweeteners -- deliberately thin -----------------------
    ("SUGAR1",     "Sugar 1kg",                         "pack",   "48.00",  "43.00",  "Sugar & Salt"),
    ("SALT1",      "Tata Salt 1kg",                     "pack",   "28.00",  "25.00",  "Sugar & Salt"),
    ("JAGGERY1",   "Gud Jaggery 1kg",                   "pack",   "78.00",  "60.00",  "Sugar & Salt"),

    # --- spices -------------------------------------------------------------
    ("HALDI200",   "Everest Haldi Powder 200g",         "pack",   "82.00",  "62.00",  "Spices"),
    ("MIRCH200",   "Everest Red Chilli Powder 200g",    "pack",   "105.00", "79.00",  "Spices"),
    ("DHANIA200",  "Everest Coriander Powder 200g",     "pack",   "88.00",  "66.00",  "Spices"),
    ("GARAM100",   "Everest Garam Masala 100g",         "pack",   "92.00",  "68.00",  "Spices"),
    ("JEERA200",   "Cumin Jeera Seeds 200g",            "pack",   "148.00", "111.00", "Spices"),
    ("RAI100",     "Mustard Rai Seeds 100g",            "pack",   "32.00",  "24.00",  "Spices"),
    ("HING50",     "Hing Asafoetida 50g",               "box",    "118.00", "88.00",  "Spices"),

    # --- tea and coffee -----------------------------------------------------
    ("TEA250",     "Tata Tea Gold 250g",                "pack",   "190.00", "135.00", "Tea & Coffee"),
    ("TEA1K",      "Tata Tea Premium 1kg",              "pack",   "560.00", "420.00", "Tea & Coffee"),
    ("REDLBL500",  "Brooke Bond Red Label 500g",        "pack",   "285.00", "214.00", "Tea & Coffee"),
    ("COFFEE50",   "Nescafe Classic Coffee 50g",        "jar",    "195.00", "146.00", "Tea & Coffee"),
    ("BRU100",     "Bru Instant Coffee 100g",           "jar",    "330.00", "247.00", "Tea & Coffee"),

    # --- packaged dairy -----------------------------------------------------
    ("MILKPOW500", "Amul Milk Powder 500g",             "pack",   "265.00", "212.00", "Dairy"),
    ("BUTTER100",  "Amul Butter 100g",                  "pack",   "58.00",  "50.00",  "Dairy"),
    ("CHEESE200",  "Amul Cheese Slices 200g",           "pack",   "135.00", "112.00", "Dairy"),
    ("CONDMILK",   "Milkmaid Condensed Milk 400g",      "tin",    "145.00", "112.00", "Dairy"),

    # --- biscuits -----------------------------------------------------------
    ("PARLEG",     "Parle-G Biscuits 800g",             "pack",   "85.00",  "66.00",  "Biscuits"),
    ("MARIE250",   "Britannia Marie Gold 250g",         "pack",   "42.00",  "32.00",  "Biscuits"),
    ("GOODDAY",    "Britannia Good Day Cashew 200g",    "pack",   "55.00",  "42.00",  "Biscuits"),
    ("OREO120",    "Cadbury Oreo Biscuits 120g",        "pack",   "40.00",  "30.00",  "Biscuits"),
    ("HIDESEEK",   "Parle Hide and Seek 200g",          "pack",   "60.00",  "46.00",  "Biscuits"),

    # --- namkeen and chips --------------------------------------------------
    ("BHUJIA200",  "Haldiram Bhujia 200g",              "pack",   "55.00",  "41.00",  "Namkeen"),
    ("ALOOBHU400", "Haldiram Aloo Bhujia 400g",         "pack",   "105.00", "79.00",  "Namkeen"),
    ("LAYS52",     "Lays Classic Salted 52g",           "pack",   "20.00",  "15.00",  "Namkeen"),
    ("KURKURE90",  "Kurkure Masala Munch 90g",          "pack",   "20.00",  "15.00",  "Namkeen"),

    # --- noodles and ready to cook ------------------------------------------
    ("MAGGI280",   "Maggi Noodles 4 pack 280g",         "pack",   "56.00",  "45.00",  "Noodles & Pasta"),
    ("PASTA500",   "Del Monte Penne Pasta 500g",        "pack",   "105.00", "79.00",  "Noodles & Pasta"),
    ("SEVIYAN450", "Bambino Vermicelli 450g",           "pack",   "48.00",  "37.00",  "Noodles & Pasta"),

    # --- beverages ----------------------------------------------------------
    ("HORLICKS500", "Horlicks Health Drink 500g",       "jar",    "285.00", "228.00", "Beverages"),
    ("BOURNVITA500", "Cadbury Bournvita 500g",          "jar",    "245.00", "196.00", "Beverages"),
    ("FROOTI600",  "Frooti Mango Drink 600ml",          "bottle", "45.00",  "36.00",  "Beverages"),
    ("COKE750",    "Coca-Cola 750ml",                   "bottle", "45.00",  "37.00",  "Beverages"),

    # --- personal care ------------------------------------------------------
    ("LIFEBUOY4",  "Lifebuoy Soap 4 x 125g",            "pack",   "152.00", "112.00", "Personal Care"),
    ("DOVE100",    "Dove Beauty Bar 100g",              "pack",   "68.00",  "50.00",  "Personal Care"),
    ("CLINIC340",  "Clinic Plus Shampoo 340ml",         "bottle", "245.00", "178.00", "Personal Care"),
    ("COLGATE200", "Colgate Strong Teeth 200g",         "tube",   "115.00", "84.00",  "Personal Care"),
    ("HEADSH180",  "Head and Shoulders Shampoo 180ml",  "bottle", "215.00", "157.00", "Personal Care"),
    ("GILLETTE",   "Gillette Guard Razor",              "pc",     "35.00",  "25.00",  "Personal Care"),

    # --- home care ----------------------------------------------------------
    ("SURF1K",     "Surf Excel Easy Wash 1kg",          "pack",   "135.00", "102.00", "Home Care"),
    ("ARIEL1K",    "Ariel Matic Front Load 1kg",        "pack",   "245.00", "184.00", "Home Care"),
    ("VIM500",     "Vim Dishwash Gel 500ml",            "bottle", "115.00", "86.00",  "Home Care"),
    ("LIZOL975",   "Lizol Floor Cleaner 975ml",         "bottle", "205.00", "154.00", "Home Care"),
    ("HARPIC500",  "Harpic Toilet Cleaner 500ml",       "bottle", "98.00",  "73.00",  "Home Care"),

    # --- baby ---------------------------------------------------------------
    ("PAMPERS9",   "Pampers Baby Pants Medium 9s",      "pack",   "199.00", "155.00", "Baby Care"),
    ("CERELAC300", "Nestle Cerelac Wheat 300g",         "pack",   "285.00", "222.00", "Baby Care"),

    # --- dry fruits ---------------------------------------------------------
    ("ALMOND250",  "Almonds Badam 250g",                "pack",   "285.00", "214.00", "Dry Fruits"),
    ("CASHEW250",  "Cashews Kaju 250g",                 "pack",   "325.00", "244.00", "Dry Fruits"),
    ("RAISIN250",  "Raisins Kishmish 250g",             "pack",   "125.00", "94.00",  "Dry Fruits"),
]

#: The columns the importer maps. `category` is extra and ignored by the
#: parser -- it is there so a shopkeeper can sort the sheet and build shelves
#: from it. `scripts/gen_catalog_sheet.py --check` proves it is not mistaken
#: for one of the five it does read.
HEADERS = ["sku", "name", "unit", "price", "cost", "category"]

INK = "1E2A36"
ACCENT = "496580"
HAIRLINE = "E2E8F0"


def margin_bps(price: str, cost: str) -> int:
    """The gate's own arithmetic, in integers, on the sale price."""
    p, c = int(round(float(price) * 100)), int(round(float(cost) * 100))
    return (p - c) * 10_000 // p


def headroom_bps(price: str, cost: str, floor: int) -> int:
    """Largest discount that still clears `floor`, in bps. -1 when none does.

    Delegates to the GATE'S OWN function rather than the closed form. The
    closed form disagrees: on RAI100 it says 14.77% where the binary search
    allows 14.75%, because it rounds the wrong way across an integer-paise
    boundary. Two bps is nothing to a shopkeeper and everything to this file --
    a reference table that promises a discount the gate then refuses is worse
    than no table, and the only way to be certain it never does is to ask the
    same function the gate asks.
    """
    return max_discount_for_margin(
        int(round(float(price) * 100)), int(round(float(cost) * 100)), floor
    )


def _style_header(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor=ACCENT)
    edge = Side(style="thin", color=HAIRLINE)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(bottom=edge)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def build_products(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Products"
    ws.append(HEADERS)

    for sku, name, unit, price, cost, category in PRODUCTS:
        ws.append([sku, name, unit, float(price), float(cost), category])

    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="right")

    for col, width in zip("ABCDEF", (14, 38, 10, 11, 11, 18)):
        ws.column_dimensions[col].width = width
    _style_header(ws, len(HEADERS))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(PRODUCTS) + 1}"


def build_guide(wb: Workbook) -> None:
    """Everything the Products sheet must not say.

    Never parsed -- the importer reads sheet one only -- so this can carry
    prose, a worked example and a margin table without any chance of a stray
    column being read as a price.
    """
    ws = wb.create_sheet("Read me")
    bold = Font(bold=True, color=INK)
    head = Font(bold=True, size=13, color=INK)

    def say(text: str = "", *, style: Font | None = None) -> None:
        ws.append([text])
        if style:
            ws.cell(row=ws.max_row, column=1).font = style

    say("How to load this catalogue", style=head)
    say()
    say("1. Merchant console -> Products.")
    say("2. Drop this file on the upload area. Nothing is written yet.")
    say("3. Check the preview: it shows every row, the columns it matched, and")
    say("   any row it refuses, with the reason.")
    say("4. Tick 'Replace the current catalogue', then Import.")
    say()
    say("What Replace does", style=bold)
    say("Products in this sheet are added or updated. Products NOT in this sheet")
    say("are retired -- marked inactive, never deleted. Deleting is not offered")
    say("because shelves and past conversations reference products by code, and")
    say("a delete would silently empty a shelf and break an old audit trail.")
    say("Re-importing a retired code brings it straight back.")
    say()
    say("The six original demo codes -- ATTA5, RICE5, DAL1K, OIL1L, SUGAR1,")
    say("TEA250 -- are kept on purpose, so existing shelves and live campaigns")
    say("keep working after a Replace.")
    say()

    say("The columns", style=head)
    say()
    for label, text in [
        ("sku", "Your item code. Any format. Uppercased on import; must be unique."),
        ("name", "What the shopper is shown, and what the assistant says out loud."),
        ("unit", "bag, pack, bottle, tin, pc. Free text, 16 characters."),
        ("price", "Shelf price in rupees. At least 1.00."),
        ("cost", "What you pay for it. REQUIRED, and must be below price."),
        ("category", "Ignored by the importer. Here so you can sort and build shelves."),
    ]:
        ws.append([label, text])
        ws.cell(row=ws.max_row, column=1).font = bold
    say()
    say("Column names are matched loosely -- MRP, Rate, Selling Price, Cost")
    say("Price and Purchase Price are all understood. You do not have to use")
    say("these exact words in your own sheet.")
    say()
    say("Cost never reaches a shopper. It is used server-side for one thing:")
    say("the margin floor, below. It is not in the customer's session payload")
    say("and it is not in the assistant's context window.")
    say()

    say("Why cost decides how much you can haggle", style=head)
    say()
    say("A campaign has a margin floor. The gate will not approve a discount")
    say("that takes the margin on the SALE price below it. So:")
    say()
    say("    largest discount = 1 - (1 - margin) / (1 - floor)")
    say()
    say("An item carrying a 14% margin, under a 12% floor, can give away 2.3%")
    say("and no more -- and the assistant will say 'I cannot go below cost on")
    say("this one'. That is the gate working, not a fault.")
    say()
    say("The costs here are set so most items carry 20-27%, which leaves 9-15%")
    say("of room under a 12% floor. Put your real costs in and the negotiation")
    say("narrows to match, immediately and honestly.")
    say()
    say("Sugar and salt are deliberately left thin, below a 12% floor, so the")
    say("gate refuses them outright. Butter sits just above it with 2% of room.")
    say("That is worth having: a catalogue where everything is discountable")
    say("never shows the gate saying no, and the refusal is the product.")
    say()
    say("The last column below is computed by the gate's own function, not by a")
    say("formula written twice -- so it can never promise room the gate will")
    say("then refuse.")
    say()

    start = ws.max_row + 1
    ws.append(["sku", "name", "margin", "most it can give at a 12% floor"])
    for col in range(1, 5):
        c = ws.cell(row=start, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=ACCENT)

    rows = sorted(PRODUCTS, key=lambda p: margin_bps(p[3], p[4]))
    for sku, name, _unit, price, cost, _cat in rows:
        m = margin_bps(price, cost)
        h = headroom_bps(price, cost, 1_200)
        ws.append([
            sku, name, m / 100 / 100,
            "refused - margin is under the floor" if h < 0 else h / 100 / 100,
        ])
        ws.cell(row=ws.max_row, column=3).number_format = "0.0%"
        if h >= 0:
            ws.cell(row=ws.max_row, column=4).number_format = "0.0%"

    for col, width in zip("ABCD", (14, 40, 12, 32)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = f"A{start + 1}"


def main() -> int:
    out = Path(__file__).resolve().parents[2] / "docs" / "kirana-catalog.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    build_products(wb)
    build_guide(wb)
    wb.save(out)

    codes = [p[0] for p in PRODUCTS]
    if len(set(codes)) != len(codes):
        print("DUPLICATE SKU in PRODUCTS", file=sys.stderr)
        return 1

    print(f"wrote {out} -- {len(PRODUCTS)} products")
    thin = [p[0] for p in PRODUCTS if headroom_bps(p[3], p[4], 1_200) < 0]
    print(f"deliberately un-discountable at a 12% floor: {', '.join(thin)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
