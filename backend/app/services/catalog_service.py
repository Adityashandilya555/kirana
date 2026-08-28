"""Importing a shopkeeper's spreadsheet.

The gap this closes is small to describe and the whole reason the console
exists: a shop's price list lives in Excel, and the database only accepts
integer paise with cost strictly below price. Everything here is about
crossing that gap without making the shopkeeper learn our vocabulary.

Three decisions worth stating:

  * COLUMN NAMES ARE GUESSED, NOT DEMANDED. Real sheets say "MRP", "Rate",
    "Selling Price" or "Price (Rs)". Requiring an exact header would mean the
    first thing the product does is reject the shopkeeper's own file.

  * MONEY IS PARSED AS DECIMAL, NEVER FLOAT. "1,285.50" and "₹ 1285.5" are
    the same 128550 paise. float(285.15) * 100 is 28514.999... which truncates
    to 28514, and a one-paise error at import becomes an AMOUNT_MISMATCH at
    settlement weeks later, with nothing to point at.

  * EVERY ROW IS VALIDATED AND RETURNED, GOOD OR BAD. The import is a preview
    first and a write second, so the shopkeeper sees exactly what will land and
    which rows are refused, before anything changes.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

log = logging.getLogger("kirana.catalog")

MIN_PRICE_PAISE = 100  # the schema's own floor: check (price_paise >= 100)
MAX_ROWS = 2_000

#: Cells past this in one row are dropped. A crafted sheet can declare a row
#: several million columns wide; we only ever read five of them.
MAX_COLS = 64

#: Header aliases, lower-cased and stripped of punctuation before matching.
#: Ordered by how strongly each implies the field, because "price" appears
#: inside "cost price" and would otherwise steal it.
_ALIASES: dict[str, tuple[str, ...]] = {
    "sku": ("sku", "code", "itemcode", "productcode", "item code", "product code",
            "barcode", "id", "article"),
    "name": ("name", "productname", "itemname", "product", "item", "description",
             "particulars"),
    "unit": ("unit", "uom", "pack", "packing", "packsize", "size"),
    "cost_paise": ("costprice", "cost price", "cost", "buyprice", "buy price",
                   "purchaseprice", "purchase price", "landedcost", "wholesale"),
    "price_paise": ("sellingprice", "selling price", "saleprice", "sale price",
                    "price", "mrp", "rate", "retail"),
}

_MONEY_STRIP = re.compile(r"[^\d.\-]")
_NORMALISE = re.compile(r"[^a-z0-9 ]")

#: Leading characters Excel and Sheets treat as the start of a formula. A cell
#: like `=cmd|'/c calc'!A1` in a product name is inert in our JSON API and in
#: React, which is why this is a flag rather than a rejection -- but it becomes
#: code execution on the shopkeeper's machine the moment anyone adds "export
#: catalog to CSV", which is an obvious next feature given the import exists.
#: Surfacing it at import means the row is visible before it is ever stored.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


@dataclass
class Row:
    line: int
    sku: str = ""
    name: str = ""
    unit: str = "pc"
    price_paise: int = 0
    cost_paise: int = 0
    errors: list[str] = field(default_factory=list)
    #: Notes that do not block the import, unlike `errors`.
    warnings: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors

    @property
    def margin_bps(self) -> int:
        if self.price_paise <= 0:
            return 0
        return (self.price_paise - self.cost_paise) * 10_000 // self.price_paise

    def to_item(self) -> dict[str, Any]:
        return {
            "sku": self.sku, "name": self.name, "unit": self.unit,
            "price_paise": self.price_paise, "cost_paise": self.cost_paise,
        }

    def public(self) -> dict[str, Any]:
        return {
            "line": self.line, "sku": self.sku, "name": self.name, "unit": self.unit,
            "price_paise": self.price_paise, "cost_paise": self.cost_paise,
            "margin_bps": self.margin_bps, "ok": self.ok, "errors": self.errors,
            "warnings": self.warnings,
        }


@dataclass
class ParsedSheet:
    rows: list[Row]
    mapping: dict[str, str]
    missing: list[str]

    @property
    def valid(self) -> list[Row]:
        return [r for r in self.rows if r.ok]

    @property
    def rejected(self) -> list[Row]:
        return [r for r in self.rows if not r.ok]

    def public(self) -> dict[str, Any]:
        return {
            "mapping": self.mapping,
            "missing_columns": self.missing,
            "total": len(self.rows),
            "accepted": len(self.valid),
            "rejected": len(self.rejected),
            "rows": [r.public() for r in self.rows],
        }


class SheetError(ValueError):
    """The file could not be read at all, as opposed to rows being invalid."""


def _norm(header: str) -> str:
    return _NORMALISE.sub("", str(header or "").strip().lower())


def _match_columns(headers: list[str]) -> tuple[dict[str, int], list[str]]:
    """Map our field names onto whatever the shopkeeper called their columns."""
    normalised = [_norm(h) for h in headers]
    used: set[int] = set()
    mapping: dict[str, int] = {}

    # Cost before price on purpose: "cost price" contains "price", and matching
    # price first would consume the cost column and leave cost unmapped.
    for field_name in ("sku", "name", "unit", "cost_paise", "price_paise"):
        for alias in _ALIASES[field_name]:
            key = _norm(alias)
            for idx, head in enumerate(normalised):
                if idx in used or not head:
                    continue
                if head == key or key in head:
                    mapping[field_name] = idx
                    used.add(idx)
                    break
            if field_name in mapping:
                break

    missing = [f for f in ("sku", "name", "price_paise") if f not in mapping]
    return mapping, missing


def parse_money(raw: Any) -> int | None:
    """Rupees as written by a human -> integer paise, or None if unreadable.

    Decimal throughout. The obvious int(float(x) * 100) silently loses a paise
    on values like 285.15, and that error only surfaces much later as a
    settlement amount mismatch.
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        text = str(raw)
    else:
        text = str(raw)
    text = _MONEY_STRIP.sub("", text.strip())
    if not text or text in {"-", "."}:
        return None
    try:
        rupees = Decimal(text)
    except InvalidOperation:
        return None
    if rupees < 0:
        return None
    return int((rupees * 100).quantize(Decimal("1")))


def _read_rows(filename: str, blob: bytes) -> list[list[Any]]:
    lowered = (filename or "").lower()

    if lowered.endswith(".xls"):
        raise SheetError(
            "The old .xls format is not supported. Open it in Excel or Sheets "
            "and save as .xlsx or .csv."
        )

    if lowered.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise SheetError("Excel support is unavailable on this server.") from exc
        try:
            # data_only: read the computed value of a formula, not "=B2*1.2".
            wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            # Do not surface the parser's own message: openpyxl and zipfile
            # exceptions carry archive member paths and internal state.
            log.warning("openpyxl refused a workbook: %s", exc)
            raise SheetError(
                "That file could not be opened as a spreadsheet. "
                "Re-save it as .xlsx or .csv and try again."
            ) from exc
        ws = wb[wb.sheetnames[0]]

        # Consume the generator with a running count instead of materialising
        # the sheet. An xlsx is a zip: the upload cap bounds the COMPRESSED
        # size, and a compliant 4 MB file expands to gigabytes of sheet XML at
        # ordinary ratios. read_only=True gives a streaming parser and
        # `[list(r) for r in ...]` threw that away, so the whole bomb landed in
        # memory before MAX_ROWS was ever consulted.
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            if len(rows) > MAX_ROWS:
                raise SheetError(
                    f"That sheet has more than {MAX_ROWS} rows. "
                    "Split it, or remove the blank rows below your data."
                )
            rows.append(list(row[:MAX_COLS]))
        return rows

    # Everything else is treated as delimited text. utf-8-sig strips the BOM
    # Excel writes on "CSV UTF-8", which would otherwise corrupt the first
    # header and make the sku column unmappable.
    try:
        text = blob.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = blob.decode("latin-1")
        except Exception as exc:  # noqa: BLE001
            raise SheetError("That file is not readable text.") from exc
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect)]


def parse_sheet(filename: str, blob: bytes) -> ParsedSheet:
    raw = _read_rows(filename, blob)
    raw = [r for r in raw if any(str(c).strip() for c in r if c is not None)]
    if not raw:
        raise SheetError("That file has no rows in it.")
    if len(raw) - 1 > MAX_ROWS:
        raise SheetError(f"That file has more than {MAX_ROWS} products in it.")

    headers = [str(c or "") for c in raw[0]]
    index, missing = _match_columns(headers)

    mapping = {
        field_name: headers[idx] for field_name, idx in index.items()
        if idx < len(headers)
    }
    if missing:
        return ParsedSheet(rows=[], mapping=mapping, missing=missing)

    def cell(row: list[Any], field_name: str) -> Any:
        idx = index.get(field_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    seen: set[str] = set()
    rows: list[Row] = []

    for line, raw_row in enumerate(raw[1:], start=2):
        row = Row(line=line)
        row.sku = str(cell(raw_row, "sku") or "").strip().upper()
        row.name = str(cell(raw_row, "name") or "").strip()
        row.unit = (str(cell(raw_row, "unit") or "").strip() or "pc")[:16]

        if not row.sku:
            row.errors.append("Missing item code")
        elif row.sku in seen:
            row.errors.append(f"Duplicate item code {row.sku}")
        else:
            seen.add(row.sku)

        if not row.name:
            row.errors.append("Missing product name")
        elif row.name.startswith(_FORMULA_LEAD) or row.sku.startswith(_FORMULA_LEAD):
            # Not a rejection: it is a legitimate, if odd, product name here and
            # harmless everywhere this data currently goes. Flagged so it is
            # visible before it is stored, and so a future CSV export has
            # something to escape rather than discovering this later.
            row.warnings.append(
                "Starts with a spreadsheet formula character — it will be "
                "stored as plain text."
            )

        price = parse_money(cell(raw_row, "price_paise"))
        cost = parse_money(cell(raw_row, "cost_paise"))

        if price is None:
            row.errors.append("Price is missing or not a number")
        elif price < MIN_PRICE_PAISE:
            row.errors.append("Price must be at least ₹1.00")
        else:
            row.price_paise = price

        # Cost is optional in the sheet but required by the gate: without it
        # there is no margin, and with no margin the floor cannot be applied.
        # Defaulting silently to zero would make everything look infinitely
        # discountable, so an absent cost is an error, not a default.
        if cost is None:
            row.errors.append("Cost price is missing — the margin floor needs it")
        else:
            row.cost_paise = cost

        if price is not None and cost is not None and cost >= price:
            row.errors.append("Cost is not below price — this item would sell at a loss")

        rows.append(row)

    return ParsedSheet(rows=rows, mapping=mapping, missing=[])


def template_csv() -> str:
    """A starter file, so "what should my sheet look like?" has an answer."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["sku", "name", "unit", "price", "cost"])
    w.writerow(["ATTA5", "Aashirvaad Whole Wheat Atta 5kg", "bag", "285.00", "245.00"])
    w.writerow(["TEA250", "Tata Tea Gold 250g", "pack", "190.00", "135.00"])
    w.writerow(["SUGAR1", "Sugar 1kg", "pack", "48.00", "43.00"])
    return buf.getvalue()
